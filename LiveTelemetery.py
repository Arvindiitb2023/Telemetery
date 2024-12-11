import sys
from PyQt5 import QtCore, QtWidgets, uic
from PyQt5.QtGui import QKeyEvent
from PyQt5.QtCore import Qt
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
import hello_rc
import serial
import serial.tools.list_ports
import mplcyberpunk
import matplotlib.ticker as ticker
import pandas as pd
from openpyxl import load_workbook
# Global Matplotlib settings for dark-themed graphs
plt.rcParams.update({
    'axes.facecolor': '#0f1b2a',
    'axes.edgecolor': '#0f1b2a',
    'axes.labelcolor': '#0f1b2a',
    'xtick.color': '#ffffff',
    'ytick.color': '#ffffff',
    'grid.color': '#d4dbff',
    'lines.linewidth': 2,
    'figure.facecolor': '#0f1b2a',
    'figure.edgecolor': '#0f1b2a',
    'legend.facecolor': '#3b3b3b',
    'legend.edgecolor': '#444444',
    'legend.fontsize': '9',
    'axes.grid': True,
    'axes.grid.axis': 'y',
    'grid.alpha': 0.3,
    'savefig.facecolor': '#0f1b2a',
    'savefig.edgecolor': '#0f1b2a',
    'savefig.dpi': 300,
    'font.size': 10,
    'text.color': '#ffffff',
})
class RealTimeGraphApp(QtWidgets.QMainWindow):
    def __init__(self,parent = None):
        super().__init__(parent)

        # Load the UI file
        uic.loadUi("aesthetic5.2.ui", self)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        # Setup the stacked widget and connect the buttons
        self.sw1 = self.findChild(QtWidgets.QStackedWidget, 'stackedWidget_1')
        self.sw = self.findChild(QtWidgets.QStackedWidget, 'stackedWidget')
        self.sw1.setCurrentIndex(0)
        start_button = self.findChild(QtWidgets.QPushButton, 'start')
    
        refresh_button  = self.findChild(QtWidgets.QPushButton, 'refresh')
        close_button = self.findChild(QtWidgets.QPushButton , 'close_btn')
        min_button = self.findChild(QtWidgets.QPushButton , 'min')
        max_button = self.findChild(QtWidgets.QPushButton , 'max')
        self.msg = self.findChild(QtWidgets.QTextEdit , 'message')

        #series in dict
        self.series = {}
        for i in range(1,10):
           self.series[f'series{i}'] = self.findChild(QtWidgets.QCheckBox , f'series{i}')  
        
        # Connect button actions
        self.playing = False
        start_button.clicked.connect(self.start_reading)
        refresh_button.clicked.connect(self.refresh_ports)
        
        close_button.clicked.connect(self.close)
        min_button.clicked.connect(self.showMinimized)
        max_button.clicked.connect(self.toggleMaxRestore)
          

        # Create graphs for each widget
        self.graphs = {}
        for i in range(0, 4):
            figure, canvas,toolbar = self.create_canvas(self.findChild(QtWidgets.QWidget, f'graph_{i}'))
            self.graphs[f'graph_{i}'] = {'figure': figure, 'canvas': canvas , 'toolbar':toolbar}
            self.graphs[f'graph_{i}']['canvas'].installEventFilter(self)
        for i in range(2, 4):
            figure, canvas,toolbar = self.create_canvas(self.findChild(QtWidgets.QWidget, f'graph_{i}f'))
            self.graphs[f'graph_{i}f'] = {'figure': figure, 'canvas': canvas , 'toolbar':toolbar}
            self.graphs[f'graph_{i}f']['canvas'].installEventFilter(self)
        # for i in range(2,4):
        #     self.graphs[f'graph_{i}f'] = {'figure': None, 'canvas': None}
        #     self.graphs[f'graph_{i}f']['figure'], self.graphs[f'graph_{i}f']['canvas'] = self.create_canvas(
        #      self.sw.findChild(QtWidgets.QWidget, f'graph_{i}f'))
        #     self.graphs[f'graph_{i}f']['canvas'].installEventFilter(self)
        
        #Ports and All
        self.port = self.findChild(QtWidgets.QComboBox,'port_box')
        self.file = self.findChild(QtWidgets.QLineEdit ,'file')
        self.refresh_ports()

        #Intializing Variables
        self.serial_port = None
        self.timer = None
        self.data = [[0], [0], [0], [0], [0], [0]]
        self.save = [[0], [0], [0], [0], [0], [0]]
        self.a_data = [0]
        self.index = 0
        self.max1 = self.max2 = self.max3 = False
        self.window_size = 25
        self.installEventFilter(self)
        self.update_plot()

    # Function related to SERIAL
    def refresh_ports(self):
        # Clear existing ports
        self.port.clear()
        # List all serial ports
        ports = serial.tools.list_ports.comports()
        for port in ports:
                self.port.addItem(port.device) 

    def start_reading(self):  
            
        if not self.playing:
         self.sw1.setCurrentIndex(1)
         self.sw.setCurrentIndex(0)
         port = self.port.currentText()
         if not port:
            print("Please select a port.")
            return 
         try:
        # Adjust the baud rate and timeout as necessary
            self.serial_port = serial.Serial(port, baudrate=115200, timeout=5)
            self.data = [[], [], [], [], [], []]  # Clear previous data
            self.a_data = []
            self.timer = self.startTimer(100)  # Read data every 200 ms
            print(f"Connected to {port}")   
         except serial.SerialException as e:
            print(f"Error opening serial port: {e}")
         except Exception as e:
            print(f"An unexpected error occurred: {e}")
         self.playing = True    
        else:
          self.sw1.setCurrentIndex(0)
          self.playing = False 
          self.saving_excel()
          if self.serial_port and self.serial_port.is_open:
            self.serial_port.close()
            self.killTimer(self.timer)
            print("Serial port closed")  
                       
    def timerEvent(self, event):  
      if self.serial_port and self.serial_port.is_open:
        try:
            # Read a line from the serial port
            line = self.serial_port.readline().decode('latin-1').strip()
            if line:
                # Split the line into 6 values (comma-separated)
                values = [float(x) for x in line.split(',')]              
                # Ensure there are exactly 6 values
                if len(values) == 6:
                    # Append each value to the corresponding dataset
                    for i in range(6):
                        self.data[i].append(values[i])
                        self.save[i].append(values[i])   
                    # Keep time index updated
                    self.a_data.append(self.index) 
                    # Update the plot  
                    self.update_plot()
                else:
                    print("Incorrect number of values received")
        except Exception as e:
            print(f"Error reading from serial port: {e}")

    ####### END #######
    #    
    ### FUNCTION RELATED TO GRAPHS ###
    def create_canvas(self, parent_widget):
        """Creates and embeds a Matplotlib figure in a QWidget."""
        figure = Figure()
        canvas = FigureCanvas(figure)
        toolbar = NavigationToolbar(canvas , self)
        toolbar.hide()
        layout = QtWidgets.QVBoxLayout(parent_widget)
        layout.addWidget(toolbar)
        layout.addWidget(canvas)
        return figure, canvas,toolbar
    def show_toolbar(self ,graph_name):
        toolbar = self.graphs[graph_name]['toolbar']
        if toolbar.isVisible():
            toolbar.hide()
        else:
            toolbar.show()
    def update_plot(self): 
         
       if self.index > self.window_size and len(self.data[1]) > self.window_size:
        # Display only the last 'window_size' points
        self.a_data = list(range(self.index))[-self.window_size:]
        for i in range(6):
           self.data[i] = self.data[i][-self.window_size:]

    # Update the full graph views
       if self.max1:
           self.plot_graph('graph_0', 8, self.a_data, "state", '#00BFFF', self.data[0], "temp", '#65dc98', self.data[1])
           self.plot_graph('graph_2f', 3, self.a_data, "altitude", '#ff184c', self.data[2], "temp2", '#FF69B4', self.data[3])
           self.plot_graph('graph_3f', 5, self.a_data, "gyro_x", '#1afe49', self.data[4], "gyro_z", '#ADFF2F', self.data[5])
       elif self.max2:
           self.plot_graph('graph_0', 8, self.a_data, "altitude", '#ff184c', self.data[2], "temp2", '#FF69B4', self.data[3])
           self.plot_graph('graph_2f', 5, self.a_data, "gyro_x", '#1afe49', self.data[4], "gyro_z", '#ADFF2F', self.data[5])
           self.plot_graph('graph_3f', 1, self.a_data, "state", '#00BFFF', self.data[0], "temp", '#65dc98', self.data[1])
           
       elif self.max3:
           self.plot_graph('graph_0', 8, self.a_data, "gyro_x", '#1afe49', self.data[4], "gyro_z", '#ADFF2F', self.data[5])
           self.plot_graph('graph_2f', 1, self.a_data, "state", '#00BFFF', self.data[0], "temp", '#65dc98', self.data[1])
           self.plot_graph('graph_3f', 3, self.a_data, "altitude", '#ff184c', self.data[2], "temp2", '#FF69B4', self.data[3])
       else:
        self.plot_graph('graph_1', 1, self.a_data, "state", '#00BFFF', self.data[0], "temp", '#65dc98', self.data[1])
        self.plot_graph('graph_2', 3, self.a_data, "altitude", '#ff184c', self.data[2], "temp2", '#FF69B4', self.data[3])
        self.plot_graph('graph_3', 5, self.a_data, "gyro_x", '#1afe49', self.data[4], "gyro_z", '#ADFF2F', self.data[5])

       self.index += 1
      
    def plot_graph(self, graph_name, i, x_data, name_1, color_1, y_data1, name_2=None, color_2=None, y_data2=None):
     figure = self.graphs[graph_name]['figure']
     canvas = self.graphs[graph_name]['canvas']
     ax = figure.gca()  # Get current axes
     
     ax.cla()  # Clear the plot instead of recreating

     if self.series[f'series{i}'].isChecked():
        ax.plot(x_data, y_data1, label=name_1, color=color_1)
    
     if y_data2 is not None and self.series[f'series{i+1}'].isChecked():
        ax.plot(x_data, y_data2, label=name_2, color=color_2)

    # Set axis limits
     ax.set_xlim(min(x_data), max(x_data))
     ax.set_xlabel("Time (s)", color="#ffffff", fontsize=10)
     ax.set_ylabel(f"{name_1} / {name_2 if name_2 else name_1}", color="#ffffff", fontsize=10)

    # Adjust ticks and layout
     ax.xaxis.set_major_locator(ticker.MaxNLocator(nbins=5))  # Limit the number of x-ticks
     ax.tick_params(axis='x', colors="#ffffff", labelsize=8, rotation=45)  # Smaller size and rotation for readability
     ax.tick_params(axis='y', colors="#ffffff")
     ax.figure.subplots_adjust(left=0.06, right=0.94, top=0.9, bottom=0.2)  # Add padding
     

     ax.legend(loc='center left', bbox_to_anchor=(1, 1), frameon=False)
     mplcyberpunk.make_lines_glow(ax, n_glow_lines=6, diff_linewidth=0.95, alpha_line=0.2)
     canvas.draw_idle()
    
    ### END ###

    ### FUNCTION RELATED TO BUTTONS ### 
    def show_graph(self,i):
        self.sw.setCurrentIndex(1)        
        self.max1 = (i == 1)
        self.max2 = (i == 2)
        self.max3 = (i == 3)
        self.update_plot()
    def show_full(self,i):        
        self.max1 = (i == 1)
        self.max2 = (i == 2)
        self.max3 = (i == 3)
        self.update_plot()
    def show0(self):
        self.sw.setCurrentIndex(0)
        self.max1 = self.max2 = self.max3 = False
                
    def eventFilter(self, source, event):
        # Detect mouse double-click event
         if event.type() == event.MouseButtonDblClick and event.button() == Qt.LeftButton:
        # Check which graph canvas was double-clicked

           for graph_key, graph in self.graphs.items():
            if source == graph['canvas']:
                if graph_key in ['graph_1', 'graph_2', 'graph_3']:
                    self.show_graph(int(graph_key[-1]))  # Show Graph 1, 2, or 3 page
                elif graph_key == 'graph_0':
                    self.show0()
                elif graph_key == 'graph_2f':
                    self.show_full(2 if self.max1 else 3 if self.max2 else 1)
                elif graph_key == 'graph_3f':
                    self.show_full(3 if self.max1 else 1 if self.max2 else 2)
                break
         if event.type() == event.MouseButtonPress:
            if event.button() == Qt.RightButton:        
               for graph_key, graph in self.graphs.items():
                 if source == graph['canvas']:
                   if graph_key in ['graph_1', 'graph_2', 'graph_3','graph_0']:
                    self.show_toolbar(graph_key)  # Show Graph 1, 2, or 3 page
                   break
         return super(RealTimeGraphApp, self).eventFilter(source, event)
    
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Space:
           self.start_reading()
           
    def toggleMaxRestore(self):
        if self.isMaximized():
            self.showNormal()  # Restore window size
        else:
            self.showMaximized()  # Maximize window 
    ### END ###
    
    ### FUNCTION RELATED TO DATA SAVING ###
    def saving_excel(self):
        new_df = pd.DataFrame({
          'state': self.save[0],
          'temp' : self.save[1],
          'state': self.save[2],
          'temp2' : self.save[3],
          'gyro-x': self.save[4],
          'gyro-z' : self.save[5],
              })
    
        output_file = self.file.text()+ '.xlsx'
        try:
    # Load the existing workbook
            book = load_workbook(output_file)

    # Create a Pandas Excel writer using openpyxl as the engine
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # Write the DataFrame to the existing sheet without assigning book
                new_df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row, sheet_name='Sheet1')
            
            print(f'New data appended to {output_file}')
        except FileNotFoundError:
    # If the file doesn't exist, create it and write the data
            new_df.to_excel(output_file, index=False)
            self.msg.setHtml(f"<div style='text-align: center; font-size: 18px;'>Created a new file: {output_file}</div>")
            
            print(f' Created a new file: {output_file}')





if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = RealTimeGraphApp()
    window.showMaximized()
    sys.exit(app.exec_())